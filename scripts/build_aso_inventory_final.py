#!/usr/bin/env python3
"""
Build Canonical ASO Design Inventory - Final Version

This script processes ASO sequences from multiple sources and creates
a deduplicated canonical table.
"""

import csv
import re
import sys
from pathlib import Path
from collections import defaultdict

# Check for pandas
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("WARNING: pandas not installed. Install with: pip install pandas openpyxl", file=sys.stderr)
    print("Excel files will be skipped.", file=sys.stderr)

def extract_nucleotides_from_idt(idt_str):
    """Extract nucleotide bases from IDT/MOE format string.
    
    IDT/MOE format contains modification codes like /52MOErC/, /i2MOErA/, etc.
    and standalone nucleotides like *T*T*G*T* or +A*+T*+A.
    
    Strategy: Extract ALL A, T, C, G letters from the entire string in order,
    regardless of modification annotations. This preserves full-length sequences.
    
    Example:
        Input: /52MOErT/*/i2MOErA/*/i2MOErG/*/i2MOErT/*/i2MOErG/*G*C*C*A*G*T*C*T*T*G*T*C*+A*+T*+A
        Output: TAGTGCCAGTCTTGTCAATA
    """
    if not idt_str:
        return ""
    
    # Extract ALL A, T, C, G letters (case-insensitive) from entire string
    # This preserves order and captures nucleotides from both modification codes
    # and standalone positions
    nucleotides = re.findall(r'[ATCGatcg]', str(idt_str))
    seq = ''.join(nucleotides).upper()
    
    return seq

def normalize_sequence(seq):
    """Normalize sequence: extract only ATCG letters, uppercase.
    
    For plain sequences (not IDT format), extract all ATCG letters.
    This handles formats like "ATCG-ATCG" or "ATCG ATCG" or "ATCG*ATCG".
    """
    if not seq:
        return ""
    # Extract only ATCG letters (case-insensitive), preserving order
    nucleotides = re.findall(r'[ATCGatcg]', str(seq))
    return ''.join(nucleotides).upper()

def is_valid_nucleotide_sequence(seq):
    """Check if string is a valid nucleotide sequence (ATCGN only, min 10bp).
    
    Also warns if sequence is suspiciously short (< 15 nt) for ASO designs.
    """
    if not seq or len(seq) < 10:
        return False
    
    # Sanity check: warn if sequence is very short (likely extraction error)
    if len(seq) < 15:
        print(f"WARNING: Short sequence detected ({len(seq)} nt): {seq[:30]}...", file=sys.stderr)
        print("  This may indicate incomplete extraction. Full sequence may be longer.", file=sys.stderr)
    
    return bool(re.match(r'^[ATCGN]+$', seq.upper()))

def process_csv(filepath):
    """Extract sequences from CSV file."""
    sequences = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            
            # Prioritize sequence columns (idt, sequence, etc.)
            priority_cols = [col for col in headers if any(k in col.lower() 
                           for k in ['idt', 'sequence', 'seq', 'oligo', 'nt', 'nucleotide'])]
            other_cols = [col for col in headers if col not in priority_cols]
            col_order = priority_cols + other_cols
            
            for row_num, row in enumerate(reader, start=2):
                # Get row identifier
                row_id = row.get('id', row.get('ID', row.get('well', f"row_{row_num}")))
                
                # Check columns in priority order
                found = False
                for col in col_order:
                    val = row.get(col)
                    if not val:
                        continue
                    
                    # Try IDT format extraction first (for syt1.csv)
                    seq = extract_nucleotides_from_idt(val)
                    if not seq or not is_valid_nucleotide_sequence(seq):
                        # Try normal normalization
                        seq = normalize_sequence(val)
                    
                    if is_valid_nucleotide_sequence(seq):
                        sequences.append({
                            'sequence': seq,
                            'source': Path(filepath).name,
                            'id': str(row_id),
                            'column': col
                        })
                        found = True
                        break  # Found sequence for this row, move to next
    except Exception as e:
        print(f"ERROR processing {filepath}: {e}", file=sys.stderr)
    
    return sequences

def process_excel(filepath):
    """Extract sequences from Excel file."""
    if not HAS_PANDAS:
        return []
    
    sequences = []
    try:
        # Try multiple methods to read Excel file
        df = None
        
        # Method 1: Try with openpyxl directly (read_only mode, ignores styles)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, data_only=True, read_only=True)
            # Convert to DataFrame manually
            ws = wb[wb.sheetnames[0]]
            data = []
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(cell) if cell is not None else f"col_{i}" 
                              for i, cell in enumerate(row)]
                else:
                    data.append(row)
            wb.close()
            if headers and data:
                df = pd.DataFrame(data, columns=headers[:len(data[0])] if data else headers)
        except Exception as e1:
            # Method 2: Try pandas with data_only
            try:
                df = pd.read_excel(filepath, engine='openpyxl', data_only=True)
            except Exception as e2:
                # Method 3: Try pandas without data_only
                try:
                    df = pd.read_excel(filepath, engine='openpyxl')
                except Exception as e3:
                    print(f"ERROR reading {filepath}: All methods failed", file=sys.stderr)
                    print(f"  Method 1 (openpyxl direct): {e1}", file=sys.stderr)
                    print(f"  Method 2 (pandas data_only): {e2}", file=sys.stderr)
                    print(f"  Method 3 (pandas normal): {e3}", file=sys.stderr)
                    return []
        
        if df is None or df.empty:
            print(f"WARNING: {filepath} read but contains no data", file=sys.stderr)
            return []
            
    except Exception as e:
        print(f"ERROR reading {filepath}: {e}", file=sys.stderr)
        return []
    
    # Try to identify ID/name column
    id_cols = [c for c in df.columns if any(k in str(c).lower() for k in ['id', 'name', 'aso', 'well'])]
    
    for idx, row in df.iterrows():
        # Get row identifier
        if id_cols:
            row_id = str(row.get(id_cols[0], f"row_{idx+2}"))
        else:
            row_id = f"row_{idx+2}"
        
        # Check all columns
        for col in df.columns:
            val = row.get(col)
            if pd.notna(val) and val:
                val_str = str(val)
                
                # Try IDT extraction first
                seq = extract_nucleotides_from_idt(val_str)
                if not seq or not is_valid_nucleotide_sequence(seq):
                    # Try normal normalization
                    seq = normalize_sequence(val_str)
                
                if is_valid_nucleotide_sequence(seq):
                    sequences.append({
                        'sequence': seq,
                        'source': Path(filepath).name,
                        'id': row_id,
                        'column': str(col)
                    })
                    break  # Found sequence for this row
    return sequences

def main():
    script_dir = Path(__file__).parent.parent
    all_sequences = []
    
    # Find all Excel files matching the pattern
    excel_files = list(script_dir.glob("Syt1*.xlsx"))
    
    files_to_process = [
        (script_dir / "syt1.csv", "CSV"),
    ]
    # Add all found Excel files
    for excel_file in excel_files:
        files_to_process.append((excel_file, "Excel"))
    
    print("=" * 80, file=sys.stderr)
    print("Building Canonical ASO Inventory", file=sys.stderr)
    print("=" * 80, file=sys.stderr)
    
    # Process all files
    for filepath, file_type in files_to_process:
        if not filepath.exists():
            print(f"WARNING: File not found: {filepath.name}", file=sys.stderr)
            continue
        
        print(f"\nProcessing {filepath.name} ({file_type})...", file=sys.stderr)
        
        if file_type == "CSV":
            seqs = process_csv(filepath)
        else:
            seqs = process_excel(filepath)
        
        print(f"  Found {len(seqs)} sequences", file=sys.stderr)
        all_sequences.extend(seqs)
    
    if not all_sequences:
        print("\nERROR: No valid nucleotide sequences found in any file.", file=sys.stderr)
        print("Please check that files contain ASO sequences in ATCG format.", file=sys.stderr)
        sys.exit(1)
    
    print(f"\nTotal sequences found: {len(all_sequences)}", file=sys.stderr)
    
    # Deduplicate by exact sequence
    sequence_map = defaultdict(lambda: {'sources': set(), 'ids': []})
    for s in all_sequences:
        seq = s['sequence']
        sequence_map[seq]['sources'].add(s['source'])
        sequence_map[seq]['ids'].append(f"{s['id']} ({s['column']})")
    
    unique_count = len(sequence_map)
    print(f"Unique sequences after deduplication: {unique_count}", file=sys.stderr)
    
    # Create canonical table
    canonical = []
    for idx, (seq, info) in enumerate(sorted(sequence_map.items()), start=1):
        canonical.append({
            'ASO_ID': f"SYT1_ASO_{idx:03d}",
            'Sequence_5to3': seq,
            'Length_nt': len(seq),
            'Source_File': '; '.join(sorted(info['sources'])),
            'Original_Row_ID_or_Name': '; '.join(info['ids'])
        })
    
    # Write CSV
    output_file = script_dir / "canonical_syt1_aso_table.csv"
    try:
        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=['ASO_ID', 'Sequence_5to3', 'Length_nt', 
                                                   'Source_File', 'Original_Row_ID_or_Name'])
            writer.writeheader()
            writer.writerows(canonical)
        
        # Write status file
        status_file = script_dir / "aso_inventory_status.txt"
        with open(status_file, 'w') as f:
            f.write(f"SUCCESS\n")
            f.write(f"Output: {output_file}\n")
            f.write(f"Total unique ASOs: {len(canonical)}\n")
            f.write(f"Total sequences found: {len(all_sequences)}\n")
        
        print(f"\n{'='*80}", file=sys.stderr)
        print(f"Output written to: {output_file}", file=sys.stderr)
        print(f"Total unique ASOs: {len(canonical)}", file=sys.stderr)
        print(f"{'='*80}\n", file=sys.stderr)
    except Exception as e:
        print(f"ERROR writing output: {e}", file=sys.stderr)
        sys.exit(1)
    
    # Display table (to stdout)
    print(f"{'ASO_ID':<15} {'Length':<8} {'Sources':<30} {'Sequence (first 50bp)'}")
    print("-" * 110)
    for row in canonical:
        seq_preview = row['Sequence_5to3'][:50] + "..." if len(row['Sequence_5to3']) > 50 else row['Sequence_5to3']
        sources_preview = row['Source_File'][:28] + "..." if len(row['Source_File']) > 28 else row['Source_File']
        print(f"{row['ASO_ID']:<15} {row['Length_nt']:<8} {sources_preview:<30} {seq_preview}")

if __name__ == "__main__":
    main()

